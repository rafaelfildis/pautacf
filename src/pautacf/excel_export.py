"""Gera a planilha semanal 'PAUTA DE AUDIÊNCIAS' no modelo Calmon & Freitas."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .config import EQUIPE, NOME_ESCRITORIO, SUBTITULO_PAUTA
from .models import Audiencia

STATUS_OPCOES = ["Em andamento", "CONFIRMADA", "CANCELADA", "REDESIGNADA", "ADIADA", "REALIZADA"]

COLUNAS = [
    "DATA",
    "HORÁRIO",
    "NOME DA PARTE AUTORA",
    "NOME DA PARTE RÉ",
    "NÚMERO DO PROCESSO",
    "JUÍZO / VARA",
    "RESPONSÁVEL",
    "STATUS",
    "OBSERVAÇÕES / LINK",
]
LARGURAS = [12, 14, 24, 22, 20, 32, 20, 12, 46]

AZUL_ESCURO = "1F3864"
AZUL_CLARO = "D9E2F3"
CINZA_CLARO = "F2F2F2"
BRANCO = "FFFFFF"

BORDA_FINA = Side(style="thin", color="BFBFBF")
BORDA_CELULA = Border(
    left=BORDA_FINA, right=BORDA_FINA, top=BORDA_FINA, bottom=BORDA_FINA
)


def _mesclar_e_escrever(
    ws: Worksheet, ref: str, valor, fonte: Font, preenchimento=None, alinhamento=None
):
    ws.merge_cells(ref)
    celula = ws[ref.split(":")[0]]
    celula.value = valor
    celula.font = fonte
    celula.alignment = alinhamento or Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    if preenchimento:
        for linha in ws[ref]:
            for c in linha:
                c.fill = preenchimento


def gerar_pauta_excel(
    audiencias: list[Audiencia],
    data_inicio: date,
    data_fim: date,
    caminho_saida: Path | str,
) -> Path:
    """Monta o arquivo .xlsx da pauta semanal e salva em `caminho_saida`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pauta"

    for i, largura in enumerate(LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    preenchimento_topo = PatternFill(
        "solid", fgColor=AZUL_ESCURO
    )
    fonte_titulo = Font(size=20, bold=True, color=BRANCO)
    fonte_subtitulo = Font(size=11, bold=True, color=BRANCO)

    _mesclar_e_escrever(
        ws, "A1:C3", "CF", Font(size=24, bold=True, color=BRANCO), preenchimento_topo
    )
    _mesclar_e_escrever(
        ws, "D1:I3", "PAUTA DE AUDIÊNCIAS", fonte_titulo, preenchimento_topo
    )
    _mesclar_e_escrever(
        ws,
        "A4:I4",
        f"{NOME_ESCRITORIO}  |  {SUBTITULO_PAUTA}",
        fonte_subtitulo,
        PatternFill("solid", fgColor="2E5395"),
    )

    ws.row_dimensions[5].height = 6

    _mesclar_e_escrever(
        ws,
        "A6:I6",
        "PAUTA ORGANIZADA POR DATA E HORÁRIO",
        Font(size=11, bold=True, italic=True, color=AZUL_ESCURO),
        PatternFill("solid", fgColor=CINZA_CLARO),
    )

    fonte_resumo = Font(size=10, bold=True, color=AZUL_ESCURO)
    preenchimento_resumo = PatternFill("solid", fgColor=AZUL_CLARO)
    _mesclar_e_escrever(ws, "A7:B7", "TOTAL", fonte_resumo, preenchimento_resumo)
    ws["C7"] = len(audiencias)
    ws["C7"].font = fonte_resumo
    ws["C7"].fill = preenchimento_resumo
    ws["C7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D7"] = "PERÍODO"
    ws["D7"].font = fonte_resumo
    ws["D7"].fill = preenchimento_resumo
    ws["D7"].alignment = Alignment(horizontal="center", vertical="center")
    _mesclar_e_escrever(
        ws,
        "E7:G7",
        f"{data_inicio:%d/%m/%Y} a {data_fim:%d/%m/%Y}",
        fonte_resumo,
        preenchimento_resumo,
    )
    _mesclar_e_escrever(
        ws,
        "H7:I7",
        "MODALIDADE: VIRTUAL/PRESENCIAL",
        fonte_resumo,
        preenchimento_resumo,
    )

    ws.row_dimensions[8].height = 6

    linha_cabecalho = 9
    preenchimento_cabecalho = PatternFill("solid", fgColor=AZUL_ESCURO)
    fonte_cabecalho = Font(size=10, bold=True, color=BRANCO)
    for col, titulo in enumerate(COLUNAS, start=1):
        celula = ws.cell(row=linha_cabecalho, column=col, value=titulo)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = BORDA_CELULA
    ws.row_dimensions[linha_cabecalho].height = 24

    linha = linha_cabecalho + 1
    dia_anterior = None
    cor_alternada = False
    for aud in audiencias:
        if aud.data != dia_anterior:
            cor_alternada = not cor_alternada
            dia_anterior = aud.data

        preenchimento_linha = PatternFill(
            "solid", fgColor=CINZA_CLARO if cor_alternada else BRANCO
        )
        valores = [
            aud.data_formatada,
            aud.horario_formatado,
            aud.parte_autora,
            aud.parte_re,
            aud.numero_processo,
            aud.vara,
            aud.responsavel,
            aud.status,
            aud.observacoes,
        ]
        for col, valor in enumerate(valores, start=1):
            celula = ws.cell(row=linha, column=col, value=valor)
            celula.fill = preenchimento_linha
            celula.border = BORDA_CELULA
            celula.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        linha += 1

    linha_final = linha - 1
    if linha_final >= linha_cabecalho + 1:
        _adicionar_dropdown(ws, "G", linha_cabecalho + 1, linha_final, EQUIPE)
        _adicionar_dropdown(ws, "H", linha_cabecalho + 1, linha_final, STATUS_OPCOES)

    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida


def _adicionar_dropdown(
    ws: Worksheet, coluna: str, linha_ini: int, linha_fim: int, opcoes: list[str]
) -> None:
    """Adiciona uma lista suspensa (RESPONSÁVEL/STATUS) para facilitar o preenchimento manual."""
    formula = '"' + ",".join(opcoes) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{coluna}{linha_ini}:{coluna}{linha_fim}")
