from datetime import date

from openpyxl import load_workbook

from pautacf.excel_export import gerar_pauta_excel
from pautacf.ics_parser import extrair_audiencias
from pathlib import Path

CAMINHO_EXEMPLO = Path(__file__).resolve().parent.parent / "data" / "exemplos" / "agenda_exemplo.ics"


def test_gera_arquivo_excel_com_cabecalho_e_linhas(tmp_path):
    audiencias = extrair_audiencias(CAMINHO_EXEMPLO)
    saida = tmp_path / "pauta_teste.xlsx"

    caminho = gerar_pauta_excel(audiencias, date(2026, 7, 20), date(2026, 7, 24), saida)

    assert caminho.exists()
    wb = load_workbook(caminho)
    ws = wb.active
    assert ws["D1"].value == "PAUTA DE AUDIÊNCIAS"
    assert ws["A9"].value == "DATA"
    assert ws["I9"].value == "OBSERVAÇÕES / LINK"
    # 3 audiências de exemplo -> linhas 10, 11 e 12
    assert ws["A10"].value == "20/07/2026"
    assert ws["C10"].value == "JOÃO DA SILVA TESTE"
